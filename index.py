import collections
import shutil

from openpyxl import load_workbook


def make_report(log_file_name, report_template_file_name, report_output_file_name):
        wb = load_workbook(filename = log_file_name)
        sheet = wb['log']
        shutil.copyfile(report_template_file_name, report_output_file_name)
        wb_out = load_workbook(filename = report_output_file_name)
        sheet_out = wb_out['Лист1']

        # список браузеров
        browsers_list = []
        # список товаров
        products_list = []

        # восстребованные товары у мужчин
        products_man = []
        # восстребованные товары у женщин
        products_woman = []

        for j in range(2, len(sheet['D'])+1):
                browser = sheet.cell(row=j, column=4).value
                browsers_list.append(browser)

                products_temp = sheet.cell(row=j, column=8).value
                products_temp_list = products_temp.split(',')
                for product in products_temp_list:
                        products_list.append(product.strip())
                        if(sheet.cell(row=j, column=2).value == 'м'):
                                products_man.append(product.strip())
                        elif(sheet.cell(row=j, column=2).value == 'ж'):
                                products_woman.append(product.strip())


        browser_counter = collections.Counter(browsers_list).most_common(7)
        products_counter = collections.Counter(products_list).most_common(7)

        product_man_counter = collections.Counter(products_man).most_common()
        product_woman_counter = collections.Counter(products_woman).most_common()

        browsers_dict = {}

        months_dict = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}

        for j in browser_counter:
                browsers_dict[j[0]] = months_dict.copy()

        products_dict = {}
        for j in products_counter:
                products_dict[j[0]] = months_dict.copy()

        sum_of_months_browsers = months_dict.copy()
        sum_of_months_products = months_dict.copy()

        for j in range(2, len(sheet['D'])+1):
                browser = sheet.cell(row=j, column=4).value
                month = sheet.cell(row=j, column=7).value.strftime('%m')

                products_temp = sheet.cell(row=j, column=8).value
                products_temp_list = products_temp.split(',')

                for product in products_temp_list:
                        for k, v in products_dict.items():
                                if k == product.strip():
                                        for k1, v1 in v.items():
                                                if k1 == str(month):
                                                        products_dict[k][k1] += 1
                                                        sum_of_months_products[k1] += 1

                for k, v in browsers_dict.items():
                        if k == browser:
                                for k1, v1 in v.items():
                                        if k1 == str(month):
                                                browsers_dict[k][k1] += 1
                                                sum_of_months_browsers[k1] += 1


        # записываем браузеры в файл
        i = 5
        for k, v in browsers_dict.items():
                sheet_out.cell(row=i, column=1).value = k
                j = 2
                for k1, v1 in v.items():
                        sheet_out.cell(row=i, column=j).value = v1
                        j += 1
                i += 1

        j = 2
        for v in sum_of_months_browsers.values():
                sheet_out.cell(row=12, column=j).value = v
                j += 1


        # записываем товары в файл
        i = 19
        for k, v in products_dict.items():
                sheet_out.cell(row=i, column=1).value = k
                j = 2
                for k1, v1 in v.items():
                        sheet_out.cell(row=i, column=j).value = v1
                        j += 1
                i += 1

        j = 2
        for v in sum_of_months_products.values():
                sheet_out.cell(row=26, column=j).value = v
                j += 1


        sheet_out.cell(row=31, column=2).value = product_man_counter[0][0]
        sheet_out.cell(row=32, column=2).value = product_woman_counter[0][0]
        sheet_out.cell(row=33, column=2).value = product_man_counter[-1][0]
        sheet_out.cell(row=34, column=2).value = product_woman_counter[-1][0]

        wb_out.save(report_output_file_name)

# make_report('logs.xlsx', 'report_template.xlsx', 'report.xlsx')

# Заполнить раздел “Предпочтения”, вычислив самые популярные и самые не
# востребованные товары среди мужчин и женщин. Самый популярный товар -
# товар с наибольшим количеством продаж. Самый невостребованный - с наименьшим.