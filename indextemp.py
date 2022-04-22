import collections
import string

from pprint import pprint

from openpyxl import load_workbook


wb = load_workbook(filename = 'logs.xlsx')
sheet = wb['log']

wb_out = load_workbook(filename = 'report_template.xlsx')
sheet_out = wb_out['Лист1']

# список браузеров и месяцев
browsers_and_month_list = []
# список товаров и месяцев
products_and_month_list = []
for j in range (2, len(sheet['G'])+1):
        number = f'D{j}'
        browser = sheet[number].value
        number = f'G{j}'
        month = sheet[number].value.strftime('%m')
        temp_dict = f'{browser};{month}'
        browsers_and_month_list.append(temp_dict)

        number = f'H{j}'

        products = sheet[number].value.split(',')
        for product in products:
            temp_dict = f'{product.strip()};{month}'
            products_and_month_list.append(temp_dict)

browsers_and_month_top = collections.Counter(browsers_and_month_list).most_common()


products_and_month_top = collections.Counter(products_and_month_list).most_common()
# print(products_and_month_top)

# pprint(browsers_and_month_top)

# список словарей с топ по месяцам
full_list_browsers = []
one_dict = {}
for i in browsers_and_month_top:
    browser_and_month = i[0].split(';')
    value = i[1]
    one_dict = {'item': browser_and_month[0], 'month': browser_and_month[1], 'value': value}
    full_list_browsers.append(one_dict)

# список словарей с топ по месяцам
full_list_products = []
one_dict = {}
for i in products_and_month_top:
    browser_and_month = i[0].split(';')
    value = i[1]
    one_dict = {'item': browser_and_month[0], 'month': browser_and_month[1], 'value': value}
    full_list_products.append(one_dict)
#print(len(full_list))
pprint(full_list_products)

browsers_list = []
# оставляем только браузеры
for browser in browsers_and_month_list:
    temp_browser_only = browser.split(';')
    browser_only = temp_browser_only[0]
    browsers_list.append(browser_only)

product_list = []
# оставляем только товары
for product in products_and_month_list:
    temp_browser_only = product.split(';')
    browser_only = temp_browser_only[0]
    product_list.append(browser_only)

pprint(product_list)
# print(browsers_list)

# получаем ТОП 7 браузеров
browser_counter = collections.Counter(browsers_list).most_common(7)
# получаем ТОП 7 товаров
product_counter = collections.Counter(product_list).most_common(7)

print(browser_counter)
print(1)
print(product_counter)



def write_block(start_row, product_counter, full_list_products):
    i = start_row
    sum_dict = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0}
    for browser in product_counter:
        number = f'A{i}'
        sheet_out[number] = str(browser[0])

        for popular in full_list_products:
            if popular['item'] == str(browser[0]):
                if popular['month'] == '01':
                    number = f'B{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[1] += popular['value']
                if popular['month'] == '02':
                    number = f'C{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[2] += popular['value']
                if popular['month'] == '03':
                    number = f'D{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[3] += popular['value']
                if popular['month'] == '04':
                    number = f'E{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[4] += popular['value']
                if popular['month'] == '05':
                    number = f'F{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[5] += popular['value']
                if popular['month'] == '06':
                    number = f'G{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[6] += popular['value']
                if popular['month'] == '07':
                    number = f'H{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[7] += popular['value']
                if popular['month'] == '08':
                    number = f'I{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[8] += popular['value']
                if popular['month'] == '09':
                    number = f'J{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[9] += popular['value']
                if popular['month'] == '10':
                    number = f'K{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[10] += popular['value']
                if popular['month'] == '11':
                    number = f'L{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[11] += popular['value']
                if popular['month'] == '12':
                    number = f'M{i}'
                    sheet_out[number] = popular['value']
                    sum_dict[12] += popular['value']
        i += 1
        write_sum_row(1, 12, sum_dict, i)



def write_sum_row(start_column, end_column, sum_dict, i):
    for k in range(start_column, end_column + 1):
        number = f'{string.ascii_uppercase[k]}{i}'
        sheet_out[number] = sum_dict[k]


# запись блока с браузерами
write_block(5, browser_counter, full_list_browsers)
# запись блока с товарами
write_block(19, product_counter, full_list_products)




# заполнение названиями популярных товаров

# sheet_out.cell(row=25, column=20).value = "какое-то-значение"
# в словарь добавить списки с месяцами и количеством заходов

wb_out.save('report_template.xlsx')

# Открыть книгу: wb = load_workbook(filename='input_file_name.xlsx');
# Активировать лист: sheet = wb['Лист1'];
# Записать в ячейку: sheet["A1"] = "какие-то-данные";
# Сохранить файл: wb.save('output_file_name.xlsx').

